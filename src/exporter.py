"""
exporter.py
-----------
Stage 4 — Flatten, format, and export to a two-sheet Excel workbook.

Sheet 1 — "Line Items"
    Every flat row from all processed documents.
    Rows with needs_review=True are highlighted in amber.

Sheet 2 — "Document Summary"
    One row per source_file with:
      - total_items (count of line items)
      - total_cost  (sum of line_total)
      - has_review  (True if any row in the document is flagged)
      - review_count (number of flagged rows)

Formatting applied to both sheets:
    - Frozen header row
    - Auto-adjusted column widths
    - Bold, coloured header row
    - Number formatting on currency columns
    - Alternating row banding on the summary sheet

Public API
----------
    export_to_excel(rows: list[FlatRow], output_path: str | Path) -> None
"""

from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter

from .schema import FlatRow

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------

COLOUR_HEADER_BG      = "1F3864"   # Dark navy — header row background
COLOUR_HEADER_FG      = "FFFFFF"   # White — header row text
COLOUR_REVIEW_BG      = "FFF2CC"   # Amber — needs_review rows
COLOUR_REVIEW_BORDER  = "FFD966"   # Amber border for review rows
COLOUR_BAND_ODD       = "EBF0FA"   # Light blue-grey — odd banded rows (summary)
COLOUR_BAND_EVEN      = "FFFFFF"   # White — even banded rows (summary)
COLOUR_SUMMARY_HDR_BG = "2E4A7A"   # Slightly lighter navy for summary header
COLOUR_HAS_REVIEW_BG  = "FFC7CE"   # Red — summary rows where has_review=True
COLOUR_OK_BG          = "C6EFCE"   # Green — summary rows that are clean

# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

LINE_ITEMS_COLUMNS = [
    "source_file",
    "company_name",
    "date",
    "payment_terms",
    "delivery_time",
    "offer_validity",
    "sku",
    "item_name",
    "description",
    "currency",
    "price",
    "quantity",
    "tax",
    "total_amount",
    "line_total",
    "needs_review",
    "review_reason",
]

SUMMARY_COLUMNS = [
    "source_file",
    "company_name",
    "date",
    "currency",
    "payment_terms",
    "delivery_time",
    "total_items",
    "total_cost",
    "review_count",
    "has_review",
]

# Columns to format as currency (2 decimal places)
CURRENCY_COLUMNS = {"price", "tax", "total_amount", "line_total", "total_cost"}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_header_fill(colour: str) -> PatternFill:
    return PatternFill("solid", fgColor=colour)


def _make_thin_border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _auto_width(ws, min_width: int = 10, max_width: int = 60) -> None:
    """Set column widths based on the longest cell value in each column."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 4, max_width))


def _write_header(ws, columns: list[str], header_colour: str) -> None:
    """Write and style the header row (row 1)."""
    header_fill = _make_header_fill(header_colour)
    header_font = Font(bold=True, color=COLOUR_HEADER_FG, size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = _make_thin_border()

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"   # Freeze header row


def _format_cell(cell, col_name: str, currency: str | None = None) -> None:
    """Apply number format and alignment to a data cell."""
    if col_name in CURRENCY_COLUMNS:
        if currency == "USD":
            cell.number_format = '"$"#,##0.00'
        else:
            # Default to EGP format
            cell.number_format = '"EGP" #,##0.00'
        cell.alignment = Alignment(horizontal="right")
    elif col_name in {"quantity"}:
        cell.number_format = '#,##0.##'
        cell.alignment = Alignment(horizontal="right")
    elif col_name in {"needs_review", "has_review"}:
        cell.alignment = Alignment(horizontal="center")
    else:
        cell.alignment = Alignment(horizontal="left", wrap_text=False)


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def _write_line_items_sheet(ws, df: pd.DataFrame) -> None:
    """Write the Line Items sheet."""
    _write_header(ws, LINE_ITEMS_COLUMNS, COLOUR_HEADER_BG)

    review_fill   = _make_header_fill(COLOUR_REVIEW_BG)
    review_border = Border(
        left=Side(style="thin", color=COLOUR_REVIEW_BORDER),
        right=Side(style="thin", color=COLOUR_REVIEW_BORDER),
        top=Side(style="thin", color=COLOUR_REVIEW_BORDER),
        bottom=Side(style="thin", color=COLOUR_REVIEW_BORDER),
    )
    normal_fill   = _make_header_fill(COLOUR_BAND_EVEN)
    thin_border   = _make_thin_border()

    for row_idx, (_, row_data) in enumerate(df.iterrows(), start=2):
        is_review = bool(row_data.get("needs_review", False))
        row_fill   = review_fill   if is_review else normal_fill
        row_border = review_border if is_review else thin_border
        currency   = row_data.get("currency", "EGP")

        for col_idx, col_name in enumerate(LINE_ITEMS_COLUMNS, start=1):
            if col_name == "line_total":
                value = f"=(H{row_idx}*I{row_idx})+J{row_idx}"
            else:
                value = row_data.get(col_name)
                # Convert pandas NA/NaN to None for cleaner cells
                if pd.isna(value) if not isinstance(value, (str, bool)) else False:
                    value = None
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill   = row_fill
            cell.border = row_border
            _format_cell(cell, col_name, currency)

    _auto_width(ws)


def _write_summary_sheet(ws, df: pd.DataFrame) -> None:
    """Write the Document Summary sheet."""
    _write_header(ws, SUMMARY_COLUMNS, COLOUR_SUMMARY_HDR_BG)

    ok_fill     = _make_header_fill(COLOUR_OK_BG)
    review_fill = _make_header_fill(COLOUR_HAS_REVIEW_BG)
    thin_border = _make_thin_border()
    odd_fill    = _make_header_fill(COLOUR_BAND_ODD)
    even_fill   = _make_header_fill(COLOUR_BAND_EVEN)

    for row_idx, (_, row_data) in enumerate(df.iterrows(), start=2):
        has_review = bool(row_data.get("has_review", False))
        currency   = row_data.get("currency", "EGP")

        if has_review:
            row_fill = review_fill
        elif row_idx % 2 == 0:
            row_fill = even_fill
        else:
            row_fill = odd_fill

        for col_idx, col_name in enumerate(SUMMARY_COLUMNS, start=1):
            if col_name == "total_cost":
                value = f"=SUMIF('Line Items'!A:A, A{row_idx}, 'Line Items'!L:L)"
            else:
                value = row_data.get(col_name)
                if not isinstance(value, (str, bool)) and pd.isna(value):
                    value = None
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill   = row_fill
            cell.border = thin_border
            _format_cell(cell, col_name, currency)

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def export_to_excel(rows: list[FlatRow], output_path: str | Path) -> None:
    """
    Export a list of validated FlatRow objects to a formatted Excel workbook.

    Args:
        rows:        Flat rows from Stage 3 (validator.py).
        output_path: Destination .xlsx file path.

    Raises:
        ValueError:  If rows is empty.
        IOError:     If the file cannot be written.
    """
    output_path = Path(output_path)

    if not rows:
        raise ValueError("No rows to export. Ensure the pipeline produced at least one result.")

    logger.info("Exporting %d row(s) to '%s'", len(rows), output_path.name)

    # --- Build main DataFrame ---
    records = [r.model_dump() for r in rows]
    df_items = pd.DataFrame(records, columns=LINE_ITEMS_COLUMNS)

    # --- Build summary DataFrame ---
    df_summary = (
        df_items
        .groupby(["source_file", "company_name", "date", "currency"], as_index=False)
        .agg(
            total_items  = ("item_name",   "count"),
            total_cost   = ("line_total",  "sum"),
            review_count = ("needs_review", "sum"),
        )
    )
    df_summary["has_review"] = df_summary["review_count"] > 0
    df_summary["total_cost"] = df_summary["total_cost"].round(2)

    # Reorder summary columns
    df_summary = df_summary[SUMMARY_COLUMNS]

    # --- Create workbook ---
    wb = Workbook()
    ws_items   = wb.active
    ws_items.title = "Line Items"
    ws_summary = wb.create_sheet("Document Summary")

    _write_line_items_sheet(ws_items,   df_items)
    _write_summary_sheet(ws_summary, df_summary)

    # --- Save ---
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    logger.info("Workbook saved: %s", output_path.resolve())
    print(f"\n✅  Workbook saved → {output_path.resolve()}")
    print(f"    Line Items sheet  : {len(df_items)} row(s)")
    print(f"    Summary sheet     : {len(df_summary)} document(s)")
    flagged = int(df_summary['review_count'].sum())
    if flagged:
        print(f"    ⚠️  {flagged} row(s) need manual review (highlighted in amber)")
